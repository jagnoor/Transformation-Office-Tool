from __future__ import annotations


# =============================================================================
# Teaching notes: tests/test_excel_roundtrip.py
#
# Tests that we can write a workbook to bytes and read it back with the same data.
# This catches Excel parsing regressions early.
# =============================================================================

# ---------------------------------------------------------------------------
# Imports (standard library, third-party packages, and local modules)
# ---------------------------------------------------------------------------
from datetime import date
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from excel_io import (
    TASK_COLUMNS,
    WORKSTREAM_COLUMNS,
    read_roadmap_excel,
    template_bytes,
    write_roadmap_excel_bytes,
)


def test_template_has_required_sheets() -> None:
    b = template_bytes()
    wb = load_workbook(BytesIO(b), data_only=True, read_only=True)
    assert set(wb.sheetnames) >= {"Settings", "Workstreams", "Tasks"}


def test_roundtrip_excel_write_and_read() -> None:
    settings = {
        "chart_title": "Roundtrip",
        "chart_subtitle": "",
        "confidentiality_label": "",
        "overall_start_date": date(2025, 1, 1),
        "overall_end_date": date(2025, 2, 15),
        "timezone": "America/Chicago",
        "week_start_day": "Mon",
        "time_granularity": "Weekly",
        "output_dpi": 300,
        "show_today_line": True,
        "today_line_date": None,
        "page_size": "A4",
        "font_family": "Calibri",
    }

    workstreams_df = pd.DataFrame(
        [
            {"workstream": "One", "order": 1, "color": "#1F77B4"},
            {"workstream": "Two", "order": 2, "color": "#FF7F0E"},
        ]
    )

    tasks_df = pd.DataFrame(
        [
            {
                "id": "T1",
                "workstream": "One",
                "title": "Task",
                "description": "",
                "start_date": date(2025, 1, 5),
                "end_date": date(2025, 1, 10),
                "status": "planned",
                "owner": "",
                "color_override": "",
                "type": "block",
                "hyperlink": "https://example.com",
            }
        ]
    )

    # Ensure required columns exist and in correct order
    for c in WORKSTREAM_COLUMNS:
        if c not in workstreams_df.columns:
            workstreams_df[c] = pd.NA
    workstreams_df = workstreams_df[WORKSTREAM_COLUMNS]

    for c in TASK_COLUMNS:
        if c not in tasks_df.columns:
            tasks_df[c] = pd.NA
    tasks_df = tasks_df[TASK_COLUMNS]

    xlsx = write_roadmap_excel_bytes(settings, workstreams_df, tasks_df)
    payload = read_roadmap_excel(xlsx)

    assert payload.settings["chart_title"] == "Roundtrip"
    assert list(payload.workstreams_df.columns) == WORKSTREAM_COLUMNS
    assert list(payload.tasks_df.columns) == TASK_COLUMNS

    # Dates should come back as python date
    assert payload.tasks_df.loc[0, "start_date"] == date(2025, 1, 5)
    assert payload.tasks_df.loc[0, "end_date"] == date(2025, 1, 10)
