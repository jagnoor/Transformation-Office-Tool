from __future__ import annotations


# =============================================================================
# Teaching notes: excel_io.py (Excel template + read/write)
#
# This module owns everything related to .xlsx files:
#   - build_template_workbook(): creates the blank template with dropdowns
#   - read_roadmap_excel(): reads an uploaded workbook into plain Python data
#   - write_roadmap_excel_bytes(): writes the current in-app data back to Excel
#
# Why keep Excel I/O in its own file?
#   - It keeps the Streamlit UI code (app.py) simpler.
#   - It makes unit tests easier (we can test Excel round-trips without UI).
# =============================================================================

# ---------------------------------------------------------------------------
# Imports (standard library, third-party packages, and local modules)
# ---------------------------------------------------------------------------
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any, Dict, Optional, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from roadmap_models import COLOR_DROPDOWN_VALUES


SETTINGS_KEYS = [
    "chart_title",
    "chart_subtitle",
    "confidentiality_label",
    "overall_start_date",
    "overall_end_date",
    "timezone",
    "week_start_day",
    "time_granularity",
    "output_dpi",
    "show_today_line",
    "today_line_date",
    "page_size",
    "font_family",
]

WORKSTREAM_COLUMNS = ["workstream", "order", "color"]

TASK_COLUMNS = [
    "id",
    "workstream",
    "title",
    "description",
    "start_date",
    "end_date",
    "status",
    "owner",
    "color_override",
    "type",
    "hyperlink",
]


@dataclass(frozen=True)
class ExcelPayload:
    settings: Dict[str, Any]
    workstreams_df: pd.DataFrame
    tasks_df: pd.DataFrame


def _coerce_date(value: Any) -> Optional[date]:
    """Convert a cell value into a Python date (or None). Handles Excel dates, datetimes, strings, and pandas timestamps."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    # pandas sometimes gives Timestamp
    if hasattr(value, "to_pydatetime"):
        try:
            dt = value.to_pydatetime()
            if isinstance(dt, datetime):
                return dt.date()
        except Exception:
            pass
    if isinstance(value, str):
        v = value.strip()
        if not v:
            return None
        # Try common formats
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d-%b-%Y", "%b %d %Y"):
            try:
                return datetime.strptime(v, fmt).date()
            except ValueError:
                continue
    return None


def build_template_workbook() -> Workbook:
    """Create the blank Roadmap_Input_TEMPLATE.xlsx workbook with the required sheets and dropdown validations."""
    wb = Workbook()
    wb.remove(wb.active)

    # Settings sheet
    ws = wb.create_sheet("Settings")
    ws.append(["key", "value"])
    header_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left")

    today = date.today()
    defaults: Dict[str, Any] = {
        "chart_title": "Roadmap",
        "chart_subtitle": "",
        "confidentiality_label": "Dayforce Confidential",
        "overall_start_date": today,
        "overall_end_date": today + timedelta(days=90),
        "timezone": "America/Chicago",
        "week_start_day": "Mon",
        "time_granularity": "Weekly",
        "output_dpi": 300,
        "show_today_line": True,
        "today_line_date": "",
        "page_size": "A3",
        "font_family": "Calibri",
    }

    for key in SETTINGS_KEYS:
        ws.append([key, defaults.get(key, "")])

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40
    ws.freeze_panes = "A2"

    # Data validations for selected settings
    dv_week_start = DataValidation(type="list", formula1='"Mon,Sun"', allow_blank=False)
    dv_gran = DataValidation(type="list", formula1='"Weekly,Monthly"', allow_blank=False)
    dv_dpi = DataValidation(type="list", formula1='"150,300,600"', allow_blank=False)
    dv_page = DataValidation(type="list", formula1='"A3,A4"', allow_blank=False)
    dv_bool = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)

    ws.add_data_validation(dv_week_start)
    ws.add_data_validation(dv_gran)
    ws.add_data_validation(dv_dpi)
    ws.add_data_validation(dv_page)
    ws.add_data_validation(dv_bool)

    # Find the row numbers of the keys and apply validation to the value cell
    key_to_row = {ws.cell(row=r, column=1).value: r for r in range(2, ws.max_row + 1)}
    dv_week_start.add(ws.cell(row=key_to_row["week_start_day"], column=2))
    dv_gran.add(ws.cell(row=key_to_row["time_granularity"], column=2))
    dv_dpi.add(ws.cell(row=key_to_row["output_dpi"], column=2))
    dv_page.add(ws.cell(row=key_to_row["page_size"], column=2))
    dv_bool.add(ws.cell(row=key_to_row["show_today_line"], column=2))

    # Date formats
    for k in ("overall_start_date", "overall_end_date", "today_line_date"):
        r = key_to_row[k]
        ws.cell(row=r, column=2).number_format = "yyyy-mm-dd"

    # Workstreams sheet
    ws_w = wb.create_sheet("Workstreams")
    ws_w.append(WORKSTREAM_COLUMNS)
    for c in ws_w[1]:
        c.font = Font(bold=True)
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left")
    ws_w.freeze_panes = "A2"
    ws_w.column_dimensions["A"].width = 26
    ws_w.column_dimensions["B"].width = 10
    ws_w.column_dimensions["C"].width = 14

    # Workstream color dropdown (English names; no hex required).
    # The app still accepts hex for backwards-compatibility, but the template keeps it simple.
    color_formula = '"' + ",".join(COLOR_DROPDOWN_VALUES) + '"'
    dv_ws_color = DataValidation(type="list", formula1=color_formula, allow_blank=True)
    ws_w.add_data_validation(dv_ws_color)
    dv_ws_color.add("C2:C1000")

    # Tasks sheet
    ws_t = wb.create_sheet("Tasks")
    ws_t.append(TASK_COLUMNS)
    for c in ws_t[1]:
        c.font = Font(bold=True)
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left")
    ws_t.freeze_panes = "A2"
    col_widths = {
        "A": 14,  # id
        "B": 22,  # workstream
        "C": 30,  # title
        "D": 40,  # description
        "E": 14,  # start
        "F": 14,  # end
        "G": 14,  # status
        "H": 18,  # owner
        "I": 16,  # color_override
        "J": 12,  # type
        "K": 38,  # hyperlink
    }
    for col, w in col_widths.items():
        ws_t.column_dimensions[col].width = w

    # Data validations for tasks
    dv_status = DataValidation(type="list", formula1='"planned,in_progress,done,risk"', allow_blank=True)
    dv_type = DataValidation(type="list", formula1='"block,milestone"', allow_blank=True)
    dv_color = DataValidation(type="list", formula1=color_formula, allow_blank=True)
    ws_t.add_data_validation(dv_status)
    ws_t.add_data_validation(dv_type)
    ws_t.add_data_validation(dv_color)
    # Apply to a reasonable range (users can extend)
    dv_status.add("G2:G1000")
    dv_type.add("J2:J1000")
    dv_color.add("I2:I1000")

    # Date formats for tasks
    for cell_range in ("E2:E1000", "F2:F1000"):
        for row in ws_t[cell_range]:
            for cell in row:
                cell.number_format = "yyyy-mm-dd"

    return wb


def template_bytes() -> bytes:
    """Return the template workbook as raw .xlsx bytes (ready for a download button)."""
    wb = build_template_workbook()
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def write_roadmap_excel_bytes(
    settings: Dict[str, Any],
    workstreams_df: pd.DataFrame,
    tasks_df: pd.DataFrame,
) -> bytes:
    """Serialize the current (possibly edited) data back into an .xlsx workbook.

    This is intentionally tolerant of "messy" UI inputs (blank strings, NaNs,
    mixed types) so non-technical users can round-trip their edits.
    """
    wb = build_template_workbook()

    def _is_blank(value: Any) -> bool:
        """True if value is None/NaN/NaT/pd.NA or an empty/whitespace string."""

        if value is None:
            return True
        try:
            if pd.isna(value):
                return True
        except Exception:
            pass
        if isinstance(value, str):
            return value.strip() == ""
        return False

    # -----------------
    # Settings
    # -----------------
    ws = wb["Settings"]
    key_to_row = {ws.cell(row=r, column=1).value: r for r in range(2, ws.max_row + 1)}

    for k in SETTINGS_KEYS:
        if k not in key_to_row:
            continue
        r = key_to_row[k]
        v = settings.get(k)

        # Keep blanks truly blank in Excel
        if _is_blank(v):
            ws.cell(row=r, column=2, value=None)
            continue

        # Coerce date-like
        if k in {"overall_start_date", "overall_end_date", "today_line_date"}:
            ws.cell(row=r, column=2, value=_coerce_date(v))
            ws.cell(row=r, column=2).number_format = "yyyy-mm-dd"
            continue

        ws.cell(row=r, column=2, value=v)

    # -----------------
    # Workstreams
    # -----------------
    ws_w = wb["Workstreams"]
    # Clear old rows (keep header)
    if ws_w.max_row > 1:
        ws_w.delete_rows(2, ws_w.max_row - 1)

    df_w = workstreams_df.copy() if workstreams_df is not None else pd.DataFrame(columns=WORKSTREAM_COLUMNS)
    for c in WORKSTREAM_COLUMNS:
        if c not in df_w.columns:
            df_w[c] = pd.NA
    df_w = df_w[WORKSTREAM_COLUMNS]

    for _, row in df_w.iterrows():
        ws_name = row.get("workstream")
        if _is_blank(ws_name):
            continue
        ws_name_s = str(ws_name).strip()
        if not ws_name_s:
            continue

        order_v = row.get("order")
        order_out = None
        if not _is_blank(order_v):
            try:
                order_out = int(order_v)
            except Exception:
                order_out = None

        color_v = row.get("color")
        color_out = None
        if not _is_blank(color_v):
            color_out = str(color_v).strip() or None

        ws_w.append([ws_name_s, order_out, color_out])

    # -----------------
    # Tasks
    # -----------------
    ws_t = wb["Tasks"]
    if ws_t.max_row > 1:
        ws_t.delete_rows(2, ws_t.max_row - 1)

    df_t = tasks_df.copy() if tasks_df is not None else pd.DataFrame(columns=TASK_COLUMNS)
    for c in TASK_COLUMNS:
        if c not in df_t.columns:
            df_t[c] = pd.NA
    df_t = df_t[TASK_COLUMNS]

    for _, row in df_t.iterrows():
        # Keep rows that have at least a workstream/title/date/id value
        if all(_is_blank(row.get(c)) for c in ("id", "workstream", "title", "start_date", "end_date")):
            continue

        out_row = []
        for c in TASK_COLUMNS:
            v = row.get(c)
            if _is_blank(v):
                out_row.append(None)
                continue
            if c in {"start_date", "end_date"}:
                out_row.append(_coerce_date(v))
            else:
                out_row.append(str(v).strip() if isinstance(v, str) else v)

        ws_t.append(out_row)

    # Set date formats for tasks
    for r in range(2, ws_t.max_row + 1):
        ws_t.cell(row=r, column=5).number_format = "yyyy-mm-dd"  # start_date
        ws_t.cell(row=r, column=6).number_format = "yyyy-mm-dd"  # end_date

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def read_roadmap_excel(excel_bytes: bytes) -> ExcelPayload:
    """
    Reads the required three-sheet workbook.

    Returns raw settings dict plus two DataFrames (Workstreams, Tasks).
    Any downstream validation happens elsewhere so the UI can show friendly issues.
    """
    # Settings via openpyxl (preserves dates/bools nicely)
    try:
        wb = load_workbook(BytesIO(excel_bytes), data_only=True)
    except Exception as e:
        raise ValueError(f"Unable to read .xlsx file. Make sure it's an Excel workbook (.xlsx). Details: {e}") from e

    required_sheets = {"Settings", "Workstreams", "Tasks"}
    missing = required_sheets - set(wb.sheetnames)
    if missing:
        raise ValueError(f"Missing required sheet(s): {', '.join(sorted(missing))}. Expected: Settings, Workstreams, Tasks.")

    ws = wb["Settings"]
    settings: Dict[str, Any] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        key = row[0]
        val = row[1] if len(row) > 1 else None
        if key is None:
            continue
        key_s = str(key).strip()
        if not key_s:
            continue
        settings[key_s] = val

    # Workstreams/Tasks via pandas
    buf = BytesIO(excel_bytes)
    try:
        workstreams_df = pd.read_excel(buf, sheet_name="Workstreams", engine="openpyxl")
        buf.seek(0)
        tasks_df = pd.read_excel(buf, sheet_name="Tasks", engine="openpyxl")
    except Exception as e:
        raise ValueError(f"Unable to parse Workstreams/Tasks sheets. Details: {e}") from e

    # Normalize columns: ensure expected columns exist
    for col in WORKSTREAM_COLUMNS:
        if col not in workstreams_df.columns:
            workstreams_df[col] = pd.NA
    workstreams_df = workstreams_df[WORKSTREAM_COLUMNS]

    for col in TASK_COLUMNS:
        if col not in tasks_df.columns:
            tasks_df[col] = pd.NA
    tasks_df = tasks_df[TASK_COLUMNS]

    # Coerce date columns to python date where possible
    for dc in ("start_date", "end_date"):
        tasks_df[dc] = tasks_df[dc].apply(_coerce_date)

    # Coerce settings dates to date
    for k in ("overall_start_date", "overall_end_date", "today_line_date"):
        if k in settings:
            settings[k] = _coerce_date(settings.get(k))

    # Normalize booleans from Excel (TRUE/FALSE) or strings
    def _coerce_bool(v: Any) -> Optional[bool]:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return None
        if isinstance(v, bool):
            return v
        if isinstance(v, str):
            s = v.strip().lower()
            if s in {"true", "yes", "y", "1"}:
                return True
            if s in {"false", "no", "n", "0"}:
                return False
        if isinstance(v, (int, float)):
            return bool(v)
        return None

    if "show_today_line" in settings:
        settings["show_today_line"] = _coerce_bool(settings.get("show_today_line"))

    # Normalize some settings strings
    for k in ("chart_title", "chart_subtitle", "confidentiality_label", "timezone", "week_start_day", "time_granularity", "page_size", "font_family"):
        if k in settings and settings[k] is not None:
            settings[k] = str(settings[k]).strip()

    return ExcelPayload(settings=settings, workstreams_df=workstreams_df, tasks_df=tasks_df)


def write_sample_workbook(path: str) -> None:
    """
    Writes a filled-in sample that renders nicely (based on the included PPTX structure).
    """
    wb = build_template_workbook()
    s = wb["Settings"]

    # Helper: set value by key
    key_to_row = {s.cell(row=r, column=1).value: r for r in range(2, s.max_row + 1)}
    def set_setting(key: str, value: Any) -> None:
        r = key_to_row[key]
        s.cell(row=r, column=2, value=value)

    set_setting("chart_title", "Sep 2025 Roadmap (Subject to Change)")
    set_setting("chart_subtitle", "Blocked diagram / swimlane Gantt example")
    set_setting("confidentiality_label", "Dayforce Confidential")
    set_setting("overall_start_date", date(2025, 9, 15))
    set_setting("overall_end_date", date(2026, 2, 15))
    set_setting("timezone", "America/Chicago")
    set_setting("week_start_day", "Mon")
    set_setting("time_granularity", "Weekly")
    set_setting("output_dpi", 300)
    set_setting("show_today_line", True)
    set_setting("today_line_date", "")
    set_setting("page_size", "A3")
    set_setting("font_family", "Calibri")

    ws_w = wb["Workstreams"]
    ws_t = wb["Tasks"]

    workstreams = [
        ("Release", 1, "Blue"),
        ("Legal", 2, "Orange"),
        ("Product", 3, "Green"),
        ("Marketing", 4, "Purple"),
        ("Integration", 5, "Red"),
        ("GTM", 6, "Cyan"),
        ("Project TEST", 7, "Gray"),
    ]
    for ws_name, order, color in workstreams:
        ws_w.append([ws_name, order, color])

    tasks = [
        # Legal / deal timing
        ("L-001", "Legal", "Target Close", "Deal target close milestone", date(2025, 10, 1), date(2025, 10, 1), "planned", "Legal", "", "milestone", "https://example.com"),
        # Product diligence
        ("P-001", "Product", "Tech Diligence Deep Dive", "Deep dive with founders", date(2025, 9, 18), date(2025, 9, 18), "done", "Product", "", "milestone", ""),
        ("P-002", "Product", "Cloud & Data Diligence Review", "", date(2025, 9, 26), date(2025, 9, 26), "done", "Product", "", "milestone", ""),
        ("P-003", "Product", "Integration Plan Draft", "Design initial integration approach; capture constraints", date(2025, 9, 20), date(2025, 10, 20), "in_progress", "Product", "", "block", ""),
        # Release planning & delivery
        ("R-001", "Release", "SWP 1.0 Planning", "Design and timeline the initial iFrame integration plan", date(2025, 10, 1), date(2025, 12, 15), "in_progress", "Release", "", "block", ""),
        ("R-002", "Release", "SWP 1.0 Release", "SSO + UX integration + API data access", date(2026, 1, 1), date(2026, 1, 1), "planned", "Release", "", "milestone", ""),
        ("R-003", "Release", "SWP 2.0 Planning", "Define data + AI objectives and timeline", date(2026, 1, 15), date(2026, 2, 15), "planned", "Release", "", "block", ""),
        # Marketing / comms (with overlaps to force stacking)
        ("M-001", "Marketing", "Press Release Draft", "Draft complete and distributed for review", date(2025, 9, 29), date(2025, 10, 1), "done", "Marketing", "", "block", ""),
        ("M-002", "Marketing", "Comms Plan Build-out", "Finalize messaging + distribution plan", date(2025, 9, 29), date(2025, 10, 7), "in_progress", "Marketing", "", "block", ""),
        ("M-003", "Marketing", "Discover Activities", "Mainstage + spotlights + analyst day + booth", date(2025, 10, 7), date(2025, 10, 9), "planned", "Marketing", "", "block", ""),
        ("M-004", "Marketing", "Press Release Launch Assets", "Webpage + social + customer/partner comms", date(2025, 10, 7), date(2025, 10, 10), "planned", "Marketing", "", "block", ""),
        ("M-005", "Marketing", "Webinar 1", "Marketing webinar placeholder", date(2025, 11, 18), date(2025, 11, 18), "planned", "Marketing", "", "milestone", ""),
        ("M-006", "Marketing", "Webinar 2", "Marketing webinar placeholder", date(2025, 12, 9), date(2025, 12, 9), "planned", "Marketing", "", "milestone", ""),
        # GTM activities
        ("G-001", "GTM", "Financial Model Review", "Review model with stakeholders", date(2025, 9, 16), date(2025, 9, 16), "done", "GTM", "", "milestone", ""),
        ("G-002", "GTM", "Offering Strategy Confirmed", "", date(2025, 9, 29), date(2025, 9, 29), "done", "GTM", "", "milestone", ""),
        ("G-003", "GTM", "Pricing & Packaging Build", "Build complete + contract requirements", date(2025, 10, 6), date(2025, 10, 6), "in_progress", "GTM", "", "milestone", ""),
        ("G-004", "GTM", "SFDC & QTO Systems Release", "Salesforce off-cycle release (tentative)", date(2025, 11, 21), date(2025, 11, 21), "planned", "GTM", "", "milestone", ""),
        # Integration activities
        ("I-001", "Integration", "Demo Instance Established", "Standalone data set in demo environment", date(2025, 9, 26), date(2025, 9, 26), "done", "Integration", "", "milestone", ""),
        ("I-002", "Integration", "Demo Prep", "Branding + dataset finalized; demo script", date(2025, 10, 3), date(2025, 10, 3), "done", "Integration", "", "milestone", ""),
        # Project test overlap
        ("T-001", "Project TEST", "Parallel Workstream Test A", "Overlaps with Test B to demonstrate stacking", date(2025, 10, 1), date(2025, 10, 20), "planned", "PMO", "", "block", ""),
        ("T-002", "Project TEST", "Parallel Workstream Test B", "Overlaps with Test A", date(2025, 10, 10), date(2025, 10, 25), "planned", "PMO", "", "block", ""),
        ("T-003", "Project TEST", "Parallel Workstream Test C", "Fits after A ends", date(2025, 10, 21), date(2025, 11, 5), "planned", "PMO", "", "block", ""),
    ]

    for row in tasks:
        ws_t.append(list(row))

    wb.save(path)


def write_template_file(path: str) -> None:
    """Developer helper: write the blank template workbook to disk."""
    wb = build_template_workbook()
    wb.save(path)
