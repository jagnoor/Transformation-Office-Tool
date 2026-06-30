"""Excel input/output — dead simple format for non-technical users."""
import io
from datetime import date, datetime
from typing import Tuple, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from models import WorkItem, ChartConfig
from sample_data import get_sample_rows


# ── Reading Excel ────────────────────────────────────────────────────────────

_DATE_FORMATS = ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%d/%m/%Y", "%m/%d/%y", "%Y/%m/%d")


def _parse_date(val) -> date:
    """Parse a date from various formats."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        raise ValueError("Date is required")
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(
        f"Cannot parse date '{val}'. Use YYYY-MM-DD, MM/DD/YYYY, or DD/MM/YYYY."
    )


def read_excel(file) -> Tuple[List[WorkItem], ChartConfig, List[str]]:
    """Read work items from an Excel file.

    Expects a single sheet with columns:
      Title (required), Start Date (required), End Date (required),
      Category (optional), Description (optional), Status (optional),
      Owner (optional), Label (optional)

    Returns (items, config) where config has auto-detected settings.
    """
    df = pd.read_excel(file, sheet_name=0, engine="openpyxl")

    # Normalize column names
    col_map = {}
    for col in df.columns:
        normalized = str(col).strip().lower().replace(" ", "_").replace("-", "_")
        col_map[col] = normalized
    df = df.rename(columns=col_map)

    # Find required columns with flexible matching
    title_col = _find_col(df, ["title", "name", "task", "item", "work_item", "deliverable"])
    start_col = _find_col(df, ["start_date", "start", "begin", "from", "begin_date"])
    end_col = _find_col(df, ["end_date", "end", "finish", "to", "due", "due_date", "finish_date"])

    found_cols = ", ".join(f"'{c}'" for c in df.columns) or "(no columns found — is row 1 the header?)"
    if not title_col:
        raise ValueError(
            f"Missing required column: 'Title' (or 'Name', 'Task', 'Item'). "
            f"Columns found in your file: {found_cols}"
        )
    if not start_col:
        raise ValueError(
            f"Missing required column: 'Start Date' (or 'Start', 'Begin'). "
            f"Columns found in your file: {found_cols}"
        )
    if not end_col:
        raise ValueError(
            f"Missing required column: 'End Date' (or 'End', 'Finish', 'Due'). "
            f"Columns found in your file: {found_cols}"
        )

    # Optional columns
    cat_col = _find_col(df, ["category", "workstream", "team", "group", "stream", "department"])
    desc_col = _find_col(df, ["description", "desc", "details", "notes", "note"])
    status_col = _find_col(df, ["status", "state", "progress"])
    owner_col = _find_col(df, ["owner", "assigned", "assignee", "responsible", "lead"])
    label_col = _find_col(df, ["label", "tag", "milestone_label", "id"])

    items = []
    warnings = []
    for idx, row in df.iterrows():
        title = str(row.get(title_col, "")).strip()
        if not title or title == "nan":
            continue
        try:
            start = _parse_date(row.get(start_col))
            end = _parse_date(row.get(end_col))
        except (ValueError, TypeError) as e:
            warnings.append(f"Row {idx + 2}: {e} — skipping")
            continue

        if end < start:
            start, end = end, start
            warnings.append(f"Row {idx + 2}: Start/End dates swapped for '{title}'")

        category = _safe_str(row.get(cat_col)) if cat_col else "General"
        description = _safe_str(row.get(desc_col)) if desc_col else ""
        raw_status = _safe_str(row.get(status_col)) if status_col else ""
        status = raw_status.lower().replace(" ", "_") if raw_status else "planned"
        owner = _safe_str(row.get(owner_col)) if owner_col else ""
        label = _safe_str(row.get(label_col)) if label_col else ""

        if not category:
            category = "General"
        if status not in ("planned", "in_progress", "done", "at_risk"):
            if raw_status:
                warnings.append(
                    f"Row {idx + 2}: Unrecognized status '{raw_status}' for '{title}' "
                    f"— defaulted to 'planned'"
                )
            status = "planned"

        items.append(WorkItem(
            title=title,
            start_date=start,
            end_date=end,
            category=category,
            description=description,
            status=status,
            owner=owner,
            label=label,
        ))

    # Auto-detect config from data
    config = ChartConfig()
    if items:
        config.start_date = min(it.start_date for it in items)
        config.end_date = max(it.end_date for it in items)

    return items, config, warnings


def _find_col(df, candidates):
    """Find a column matching any of the candidate names."""
    for c in candidates:
        if c in df.columns:
            return c
    return None


def _safe_str(val) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    return str(val).strip()


# ── Writing Template ─────────────────────────────────────────────────────────

def create_template_bytes(dataset_name: str = "Product Launch") -> bytes:
    """Create a beautiful, user-friendly Excel template pre-filled with sample data.

    dataset_name selects which built-in sample dataset to populate the sheet
    with — see sample_data.SAMPLE_DATASETS for available options.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Roadmap"

    # Styling
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    # Headers
    headers = [
        ("Title", 35),
        ("Start Date", 15),
        ("End Date", 15),
        ("Category", 20),
        ("Description", 40),
        ("Status", 15),
        ("Owner", 20),
        ("Label", 15),
    ]

    for col_idx, (name, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Status dropdown validation
    status_dv = DataValidation(
        type="list",
        formula1='"planned,in_progress,done,at_risk"',
        allow_blank=True,
    )
    status_dv.error = "Please select: planned, in_progress, done, or at_risk"
    status_dv.errorTitle = "Invalid Status"
    ws.add_data_validation(status_dv)
    status_dv.add(f"F2:F200")

    # Sample data sourced from the shared dataset registry (sample_data.py) so
    # the downloadable template and the in-app "Load Sample Data" flow stay in sync.
    sample_data = get_sample_rows(dataset_name)

    data_font = Font(name="Calibri", size=11)
    data_align = Alignment(vertical="center", wrap_text=False)
    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Freeze header row
    ws.freeze_panes = "A2"
    ws.sheet_properties.tabColor = "2563EB"

    # Instructions sheet
    ws_info = wb.create_sheet("Instructions")
    ws_info.sheet_properties.tabColor = "10B981"

    instructions = [
        ("Block & Gantt Creator — Input Guide", ""),
        ("", ""),
        ("Required Columns:", ""),
        ("Title", "Name of the work item (e.g., 'API Gateway Development')"),
        ("Start Date", "When work begins (YYYY-MM-DD format)"),
        ("End Date", "When work ends (YYYY-MM-DD format)"),
        ("", ""),
        ("Optional Columns:", ""),
        ("Category", "Team or workstream (e.g., Platform, Product, Security)"),
        ("Description", "Brief description of the work"),
        ("Status", "planned, in_progress, done, or at_risk"),
        ("Owner", "Person responsible"),
        ("Label", "Short label (e.g., D1, MVP, Phase 2)"),
        ("", ""),
        ("Tips:", ""),
        ("", "• Categories are used to color-code items and group them in swim lanes"),
        ("", "• For milestones, set Start Date = End Date"),
        ("", "• The app auto-detects date ranges — no settings sheet needed"),
        ("", "• Column names are flexible: 'Name' works for 'Title', 'Team' for 'Category', etc."),
        ("", ""),
        ("Block Diagram tips:", ""),
        ("", "• Use the Label column for short identifiers (e.g., 'D1', 'MVP', 'Phase 2')"),
        ("", "• Descriptions are shown inside blocks when there is enough space"),
        ("", "• Overlapping date ranges stack vertically to show parallel work"),
    ]

    title_font = Font(name="Calibri", bold=True, size=16, color="1E293B")
    section_font = Font(name="Calibri", bold=True, size=12, color="2563EB")
    key_font = Font(name="Calibri", bold=True, size=11, color="334155")
    val_font = Font(name="Calibri", size=11, color="64748B")

    for row_idx, (key, val) in enumerate(instructions, 1):
        c1 = ws_info.cell(row=row_idx, column=1, value=key)
        c2 = ws_info.cell(row=row_idx, column=2, value=val)
        if row_idx == 1:
            c1.font = title_font
        elif key and not val:
            c1.font = section_font
        elif key and val:
            c1.font = key_font
            c2.font = val_font
        else:
            c2.font = val_font

    ws_info.column_dimensions["A"].width = 25
    ws_info.column_dimensions["B"].width = 70

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def create_sample_bytes() -> bytes:
    """Create a sample Excel with realistic data (same as template's sample data)."""
    return create_template_bytes()


def write_excel(items: List[WorkItem]) -> bytes:
    """Write a list of WorkItems to an Excel file in the same format as the template.

    Used to round-trip data after in-app editing.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Roadmap"

    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    headers = [
        ("Title", 35),
        ("Start Date", 15),
        ("End Date", 15),
        ("Category", 20),
        ("Description", 40),
        ("Status", 15),
        ("Owner", 20),
        ("Label", 15),
    ]

    for col_idx, (name, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    last_row = max(2, len(items) + 1)
    status_dv = DataValidation(
        type="list",
        formula1='"planned,in_progress,done,at_risk"',
        allow_blank=True,
    )
    ws.add_data_validation(status_dv)
    status_dv.add(f"F2:F{last_row}")

    data_font = Font(name="Calibri", size=11)
    data_align = Alignment(vertical="center", wrap_text=False)
    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    for row_idx, item in enumerate(items, 2):
        row_data = [
            item.title,
            item.start_date,
            item.end_date,
            item.category,
            item.description,
            item.status,
            item.owner,
            item.label,
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    ws.freeze_panes = "A2"
    ws.sheet_properties.tabColor = "2563EB"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
